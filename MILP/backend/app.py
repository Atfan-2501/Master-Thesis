from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from pathlib import Path
import pandas as pd
import subprocess
from openpyxl import load_workbook

app = FastAPI(title="Intermodal MILP API")

# Allow your Vite dev server to call FastAPI
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---- Paths (adjust if your backend folder differs) ----
ROOT = Path(__file__).resolve().parents[1]  # .../MILP
DEFAULT_MASTER = ROOT  / "model_output" / "master_solution_decisions.xlsx"
DEFAULT_TS = ROOT  / "model_output" / "TS_Subproblem_Solution.xlsx"

# ---- Terminal name -> code mapping (matches your frontend TERMINALS keys) ----
NAME_TO_CODE = {
    "Aarau": "AARA",
    "Basel": "BASE",
    "Baselwolf": "BASE",   # your file uses Baselwolf
    "Bern": "BERN",
    "Chiasso": "CHIA",
    "Stabio": "STAB",
    "Visp": "VISP",
    "Zurich": "ZURI",
    "Lausanne": "LAUS",
    "Geneva": "GENE",
    "Luzern": "LUZE",
    "Olten": "OLTE",
    "Schaffhausen": "SCHA",
    "Winterthur": "WINT",
    "Fribourg": "FRIB",
}


def to_code(name: str) -> str:
    """Convert terminal name to code; fallback to original string if unknown."""
    if name is None:
        return ""
    return NAME_TO_CODE.get(str(name), str(name))


def parse_master(master_path: Path) -> tuple[list, dict]:
    df = pd.read_excel(master_path, sheet_name=0)

    # 2. Filter for Intermodal only and active flows
    df = df[df["mode"].str.contains("Intermodal", case=False)].copy()
    df = df[(df["TEU_y"] > 0) | (df["frequency_f"] > 0)].copy()

    # Compute revenue from price * flow (if price is per TEU)
    df["revenue_chf"] = df["price_p"].fillna(0) * df["TEU_y"].fillna(0)

    routes = []
    for _, r in df.iterrows():
        origin_name = str(r["origin"])
        dest_name = str(r["destination"])
        routes.append({
            "origin": to_code(origin_name),
            "destination": to_code(dest_name),
            "origin_name": origin_name,
            "destination_name": dest_name,
            "mode": str(r["mode"]),
            "is_open": int(r.get("x_open", 0)),
            "flow": float(r.get("TEU_y", 0.0)),
            "frequency": int(r.get("frequency_f", 0)),
            "price": float(r.get("price_p", 0.0)),
            "revenue": float(r.get("revenue_chf", 0.0)),
            "cap_per_dep": float(r.get("cap_per_dep_TEU", 0.0)),
        })

    total_flow = sum(x["flow"] for x in routes)
    total_rev = sum(x["revenue"] for x in routes)

    # Average utilization for intermodal links (if meaningful): flow / (freq * cap_per_dep)
    utilizations = []
    for x in routes:
        if x["mode"].lower().startswith("inter") and x["frequency"] > 0 and x["cap_per_dep"] > 0:
            util = 100.0 * x["flow"] / (x["frequency"] * x["cap_per_dep"])
            utilizations.append(util)

    avg_util = sum(utilizations) / len(utilizations) if utilizations else 0.0

    summary = {
        "totalFlow": total_flow,
        "totalRevenue": total_rev,
        "totalCost": 0.0,   # Phase 1: fill later from your cost breakdown sheets
        "profit": total_rev,  # Phase 1 placeholder: profit = revenue - cost
        "numRoutes": len(routes),
        "avgUtilization": avg_util,
    }
    return routes, summary

def parse_ts(ts_path: Path) -> dict:
    xls = pd.ExcelFile(ts_path)
    ts = {}

    if "OD_Summary" in xls.sheet_names:
        od = pd.read_excel(ts_path, sheet_name="OD_Summary")
        ts["od_summary"] = [
            {
                "origin": to_code(r["origin_id"]),
                "destination": to_code(r["destination_id"]),
                "origin_name": str(r["origin_id"]),
                "destination_name": str(r["destination_id"]),
                "utilization": float(r.get("utilization_%", 0.0)),
                "flow_on_trains": float(r.get("flow_on_trains", 0.0)),
                "cap_group": float(r.get("cap_group", 0.0)),
                "group_id": str(r.get("group_id", "")),
            }
            for _, r in od.iterrows()
        ]

    if "Arc_Flows" in xls.sheet_names:
        arc = pd.read_excel(ts_path, sheet_name="Arc_Flows")
        # Keep it light: only train arcs for visualization/debug
        train_arcs = arc[arc["arc_type"].astype(str).str.lower().eq("train")].copy()
        ts["train_arcs"] = [
            {
                "arc_id": str(r.get("arc_id", "")),
                "from_node": str(r.get("from_node", "")),
                "to_node": str(r.get("to_node", "")),
                "flow_TEU": float(r.get("flow_TEU", 0.0)),
                "arc_cost_per_TEU": float(r.get("arc_cost_per_TEU", 0.0)),
                "group_id": str(r.get("group_id", "")),
            }
            for _, r in train_arcs.iterrows()
        ]

    return ts

@app.get("/health")
def health():
    return {"status": "ok"}

class OperationalParams(BaseModel):
    time_multiplier: float
    price_multiplier: float
    freq_multiplier: float


@app.post("/run")
def run(params: OperationalParams):
    base_path = ROOT / "Input Data" / "master_problem_inputs_base.xlsx" 
    active_path = ROOT / "Input Data" / "master_problem_inputs_with_taste_draws.xlsx"
    
    if not base_path.exists():
        return {"error": "Base input file missing."}

    try:
        # CRITICAL: Use data_only=True to read formula results instead of strings
        wb = load_workbook(base_path, data_only=True)
        
        if "OD_Mode_Params" in wb.sheetnames:
            ws = wb["OD_Mode_Params"]
            
            for row in range(2, ws.max_row + 1):
                mode_id = ws.cell(row=row, column=3).value
                if mode_id == "Intermodal":
                    try:
                        # Values are now read as floats even if they were formulas in Excel
                        orig_t = float(ws.cell(row=row, column=4).value or 0)
                        ws.cell(row=row, column=4).value = orig_t * (1 + params.time_multiplier)
                        
                        orig_p = float(ws.cell(row=row, column=7).value or 0)
                        ws.cell(row=row, column=7).value = orig_p * (1 + params.price_multiplier)
                        
                        orig_f = float(ws.cell(row=row, column=9).value or 0)
                        ws.cell(row=row, column=9).value = orig_f * (1 + params.freq_multiplier)
                    except (ValueError, TypeError) as e:
                        print(f"Skipping row {row} due to invalid data: {e}")
            
            # Save to the active file for the solver
            wb.save(active_path)
    except Exception as e:
        print(f"Excel Update Error: {e}")

    # Trigger the solver (benders_loop.py)
    subprocess.run(["python", str(ROOT / "backend" / "benders_loop.py")], check=True)

    # 3. Read the solved outputs after the loop completes
    master_path = DEFAULT_MASTER
    ts_path = DEFAULT_TS

    if not master_path.exists():
        return {"error": f"Missing file: {master_path}"}
    if not ts_path.exists():
        return {"error": f"Missing file: {ts_path}"}

    # Use your existing parsing logic
    routes, summary = parse_master(master_path)
    ts = parse_ts(ts_path)
    

    return {
        "routes": routes,
        "summary": summary,
        "ts": ts,
    }

@app.post("/recompute-mnl")
def recompute_mnl():
    """Trigger the MNL model calculation."""
    # Define the path to your MNL script
    mnl_script = ROOT.parent / "Discrete Choice Model" / "Multinomial_Logit_Model.py"
    
    try:
        # Execute the script as a subprocess
        subprocess.run(["python", str(mnl_script)], check=True)
        return {"status": "success", "message": "MNL parameters updated."}
    except subprocess.CalledProcessError as e:
        return {"status": "error", "message": str(e)}

@app.get("/current-parameters")
def get_current_parameters():
    """Serve the latest MNL parameters to the dashboard."""
    param_path = ROOT.parent / "Discrete Choice Model" / "model_outputs" / "mnl_parameters.csv"
    if param_path.exists():
        df = pd.read_csv(param_path)
        return df.to_dict(orient="records")
    return {"error": "Parameters not yet generated."}